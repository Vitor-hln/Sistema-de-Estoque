import csv
from sqlite3 import Cursor
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta
from database import conectar
import openpyxl
from openpyxl.styles import Font
from tkinter import filedialog, messagebox
from datetime import datetime
from controllers.report_controller import ReportController

class TelaRelatorios(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        self.master = master
        
        # Título da tela
        tk.Label(self, text="Relatórios", font=("Helvetica", 16, "bold")).pack(pady=10)
        
        # Frame para filtros
        filtros_frame = tk.LabelFrame(self, text="Filtros", padx=20, pady=10)
        filtros_frame.pack(fill="x", padx=20, pady=10)
        
        # Tipo de relatório
        tk.Label(filtros_frame, text="Relatório de:").grid(row=0, column=0, sticky="w", pady=10)
        self.tipo_relatorio = ttk.Combobox(filtros_frame, width=40)
        self.tipo_relatorio['values'] = ["Movimentações de Estoque", "Peças com Estoque Baixo", 
                                       "Peças Mais Movimentadas por Entrada ou Saída","Peças Mais Movimentadas por Volume", "Valor Total em Estoque"]
        self.tipo_relatorio.current(0)
        self.tipo_relatorio.grid(row=0, column=1, sticky="w", pady=5)
        
        # Período
        tk.Label(filtros_frame, text="Período").grid(row=1, column=0, sticky="w", pady=5)
        
        periodo_frame = tk.Frame(filtros_frame)
        periodo_frame.grid(row=1, column=1, sticky="w", pady=5)
        
        tk.Label(periodo_frame, text="De:").pack(side="left", padx=(0, 5))
        self.data_inicio = tk.Entry(periodo_frame, width=12)
        self.data_inicio.pack(side="left", padx=(0, 10))
        
        tk.Label(periodo_frame, text="Até:").pack(side="left", padx=(0, 5))
        self.data_fim = tk.Entry(periodo_frame, width=12)
        self.data_fim.pack(side="left")
        
        # Botões de atalho para período
        btn_frame = tk.Frame(filtros_frame)
        btn_frame.grid(row=2, column=1, sticky="w", pady=5)
        
        self.btn_hoje = tk.Button(btn_frame, text="Hoje", command=lambda: self.definir_periodo(0))
        self.btn_hoje.pack(side="left", padx=5)
        
        self.btn_7dias = tk.Button(btn_frame, text="Últimos 7 dias", command=lambda: self.definir_periodo(7))
        self.btn_7dias.pack(side="left", padx=5)
        
        self.btn_30dias = tk.Button(btn_frame, text="Últimos 30 dias", command=lambda: self.definir_periodo(30))
        self.btn_30dias.pack(side="left", padx=5)
        
        # Botão para gerar relatório
        self.btn_gerar = tk.Button(filtros_frame, text="Gerar Relatório", command=self.gerar_relatorio,bg="#4a86e8", fg="white", width=15)
        self.btn_gerar.grid(row=1, column=2, sticky="ew", padx=60, pady=5)
        
        self.btn_exportar = tk.Button(filtros_frame, text="Exportar Relatório", command=self.exportar_para_excel, bg="#4a86e8", fg="white", width=15)
        self.btn_exportar.grid(row=2, column=2, sticky="ew", padx=60, pady=5)
        

        
        # Inicializar com o período de hoje
        self.definir_periodo(0)
        
        # Container para o relatório
        self.relatorio_frame = tk.LabelFrame(self, text="Resultado", padx=20, pady=10)
        self.relatorio_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Mostrar um relatório padrão
        self.gerar_relatorio()
    
    def definir_periodo(self, dias):
        # Define o período baseado no número de dias atrás
        hoje = datetime.now()
        data_fim = hoje.strftime("%d/%m/%Y")
        
        if dias == 0:
            data_inicio = data_fim
        else:
            data_inicio = (hoje - timedelta(days=dias)).strftime("%d/%m/%Y")
        
        self.data_inicio.delete(0, "end")
        self.data_inicio.insert(0, data_inicio)
        
        self.data_fim.delete(0, "end")
        self.data_fim.insert(0, data_fim)
    
    def gerar_relatorio(self):
        # Limpar frame de relatório
        for widget in self.relatorio_frame.winfo_children():
            widget.destroy()
            
        tipo = self.tipo_relatorio.get()
        
        # Criar tabela para relatório
        if tipo == "Movimentações de Estoque":
            self.relatorio_movimentacoes()
        elif tipo == "Peças com Estoque Baixo":
            self.relatorio_estoque_baixo()
        elif tipo == "Peças Mais Movimentadas por Entrada ou Saída":
            self.relatorio_mais_movimentados_registro()
        elif tipo == "Peças Mais Movimentadas por Volume":
            self.relatorio_mais_movimentados_volume()
        elif tipo == "Valor Total em Estoque":
            self.relatorio_valor_estoque()
        else:
            tk.Label(self.relatorio_frame, text="Selecione um tipo de relatório").pack()

    def relatorio_movimentacoes(self):
        # Criar treeview para movimento de estoque
        colunas = ("Tipo", "Produto", "Quantidade", "Usuário","Data/Hora")
        self.tabela = ttk.Treeview(self.relatorio_frame, columns=colunas, show="headings")
        
        # Definir cabeçalhos
        for col in colunas:
            self.tabela.heading(col, text=col)
            self.tabela.column(col, width=100, anchor="center")
        
        # Ajustar larguras específicas
        self.tabela.column("Data/Hora", width=150)
        self.tabela.column("Produto", width=100)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(self.relatorio_frame, orient="vertical", command=self.tabela.yview)
        self.tabela.configure(yscrollcommand=scrollbar.set)
        
        # Posicionar na tela
        self.tabela.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        try:
            # Obtenha os valores do período
            data_inicio_str = self.data_inicio.get()  # Ex: "15/03/2025"
            data_fim_str = self.data_fim.get()        # Ex: "15/03/2025"
            tipo = self.tipo_relatorio.get()

            # Converter para o formato "aaaa-mm-dd"
            data_inicio = datetime.strptime(data_inicio_str, "%d/%m/%Y").strftime("%Y-%m-%d")
            data_fim = datetime.strptime(data_fim_str, "%d/%m/%Y").strftime("%Y-%m-%d")

            controller = ReportController()
            dados = controller.obter_movimentacoes(tipo, data_inicio, data_fim)

            for mov in dados:
                data_formatada = mov['data_hora'].strftime("%d/%m/%Y %H:%M:%S")
                self.tabela.insert("", "end",
                    values=( 
                        mov['tipo'],      
                        mov['produto'],  
                        mov['quantidade'], 
                        mov['usuario'],
                        data_formatada
                    ))

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao buscar movimentações: {e}")
   
    def relatorio_estoque_baixo(self):
        controller = ReportController()

        # Criar treeview para produtos com estoque baixo
        colunas = ("Produto", "Estoque Atual", "Estoque Mínimo", "Status")
        tabela = ttk.Treeview(self.relatorio_frame, columns=colunas, show="headings")

        # Definir cabeçalhos e configurações das colunas
        for col in colunas:
            tabela.heading(col, text=col)
            tabela.column(col, width=100, anchor="center")

        tabela.column("Produto", width=250)
        tabela.column("Estoque Atual", width=100)
        tabela.column("Estoque Mínimo", width=100)
        tabela.column("Status", width=100)

        # Scrollbar
        scrollbar = ttk.Scrollbar(self.relatorio_frame, orient="vertical", command=tabela.yview)
        tabela.configure(yscrollcommand=scrollbar.set)

        # Posicionar na tela
        tabela.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        try:
            # Utiliza o controller para obter os produtos com estoque baixo
            produtos = controller.obter_produtos_estoque_baixo()
            for prod in produtos:
                status = prod.get('status', 'Desconhecido')
                tabela.insert(
                    "", "end",
                    values=(
                        prod.get('produto', 'Desconhecido'),
                        prod.get('estoque_atual', 0),
                        prod.get('estoque_minimo', 0),
                        status
                    ),
                    tags=(status.lower(),)
                )
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao buscar produtos com estoque baixo: {e}")

        # Definir cores para as tags
        tabela.tag_configure("crítico", background="#ffcccc")  # Vermelho claro para crítico
        tabela.tag_configure("baixo", background="#ffffcc")    # Amarelo claro para baixo
        tabela.tag_configure("normal", background="#ccffcc")   # Verde claro para normal
  
    def relatorio_mais_movimentados_registro(self):
        controller = ReportController()
        # Definir as colunas do Treeview
        colunas = ("Rank", "Produto", "Total Mov's", "Entradas", "Saídas")
        tabela = ttk.Treeview(self.relatorio_frame, columns=colunas, show="headings")

        # Configurar os cabeçalhos e larguras
        for col in colunas:
            tabela.heading(col, text=col)
            tabela.column(col, width=100, anchor="center")
            tabela.column("Rank", width=30)
            tabela.column("Produto", width=200)
            tabela.column("Entradas", width=15)
            tabela.column("Saídas", width=15)

        # Configurar o scrollbar
        scrollbar = ttk.Scrollbar(self.relatorio_frame, orient="vertical", command=tabela.yview)
        tabela.configure(yscrollcommand=scrollbar.set)

        # Posicionar na tela
        tabela.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        try:
            # Obtém os produtos mais movimentados
            produtos = controller.obter_mais_movimentado_registro()
            # Insere os dados na Treeview, atribuindo um ranking sequencial
            for idx, prod in enumerate(produtos, start=1):
                tabela.insert(
                    "", "end",
                    values=(
                        idx,
                        prod.get('produto', 'Desconhecido'),
                        prod.get('total_movimentacoes', 0),
                        prod.get('entradas', 0),
                        prod.get('saidas', 0)
                    )
                )
        except Exception as e:
            print("Erro ao buscar movimentações:", e)        

    def relatorio_mais_movimentados_volume(self):
        controller = ReportController()
        # Definir as colunas do Treeview
        colunas = ("Rank", "Produto", "Volume Movimentado")
        tabela = ttk.Treeview(self.relatorio_frame, columns=colunas, show="headings")

        # Configurar os cabeçalhos e larguras
        for col in colunas:
            tabela.heading(col, text=col)
            tabela.column(col, width=100, anchor="center")
            tabela.column("Rank", width=30)
            tabela.column("Produto", width=200)

        # Configurar o scrollbar
        scrollbar = ttk.Scrollbar(self.relatorio_frame, orient="vertical", command=tabela.yview)
        tabela.configure(yscrollcommand=scrollbar.set)

        # Posicionar na tela
        tabela.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        try:
            # Obtém os produtos mais movimentados
            produtos = controller.obter_mais_movimentado_volume()
            # Insere os dados na Treeview, atribuindo um ranking sequencial
            for idx, prod in enumerate(produtos, start=1):
                tabela.insert(
                    "", "end",
                    values=(
                        idx,
                        prod.get('produto', 'Desconhecido'),
                        prod.get('volume_movimentado', 0)
                    )
                )
        except Exception as e:
            print("Erro ao buscar movimentações:", e)  

    def relatorio_valor_estoque(self):
        controller= ReportController()
        # Criar treeview para valor total em estoque
        colunas = ("ID", "Produto", "Quantidade", "Valor Unitário", "Valor Total")
        self.tabela = ttk.Treeview(self.relatorio_frame, columns=colunas, show="headings")
        
        # Definir cabeçalhos
        for col in colunas:
            self.tabela.heading(col, text=col)
            self.tabela.column(col, width=100, anchor="center")
        
        # Ajustar larguras específicas
        self.tabela.column("Produto", width=250)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(self.relatorio_frame, orient="vertical", command=self.tabela.yview)
        self.tabela.configure(yscrollcommand=scrollbar.set)
        
        # Posicionar na tela
        self.tabela.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Adicionar dados de exemplo
        try:
            # Obtém os produtos mais movimentados
            valor = controller.obter_valor_total()
            # Insere os dados na Treeview, atribuindo um ranking sequencial
            for val in valor:
                self.tabela.insert("","end", values=(
                        val.get('id'),
                        val.get('produto', 'Desconhecido'),
                        val.get('quantidade', 0),
                        f"R$ {val.get('valor', 0)}",
                        f"R$ {val.get('valor_total', 0)}"
                ))
        except Exception as e:
            print("Erro ao buscar movimentações:", e)  
            
    def exportar_para_excel(self):
        if not self.tabela.get_children():
            messagebox.showwarning("Aviso", "Nenhum relatório foi gerado ainda!")
            return

        # Abrir a caixa de diálogo para salvar o arquivo
        arquivo = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Arquivo Excel", "*.   xlsx")])

        if not arquivo:
            return

        # Garantir que o nome do arquivo termina em .xlsx
        if not arquivo.endswith(".xlsx"):
            arquivo += ".xlsx"

        colunas = [self.tabela.heading(col)["text"] for col in self.tabela["columns"]]
        dados = [self.tabela.item(item, "values") for item in self.tabela.get_children()]

        try:
            # Criar um novo arquivo Excel
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Relatório"

            # Adicionar cabeçalhos
            for col_index, col_name in enumerate(colunas, start=1):
                cell = sheet.cell(row=1, column=col_index, value=col_name)
                cell.font = Font(bold=True)  # Negrito para os cabeçalhos

            # Adicionar dados na planilha
            for row_index, row_data in enumerate(dados, start=2):
                for col_index, cell_value in enumerate(row_data, start=1):
                    sheet.cell(row=row_index, column=col_index, value=cell_value)

            # Salvar o arquivo Excel corretamente
            workbook.save(arquivo)
            workbook.close()  # 🔹 Fechar corretamente para evitar corrupção do arquivo

            messagebox.showinfo("Sucesso", f"Relatório exportado com sucesso!\nSalvo em: {arquivo}")

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar para Excel: {e}")

